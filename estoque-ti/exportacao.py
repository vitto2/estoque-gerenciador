"""
Geração da planilha Excel (.xlsx) de exportação do estoque.
Mantido separado das rotas para deixar routes/equipamentos.py focado
em request/response.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

CABECALHOS = [
    "Nome/Tipo",
    "Categoria",
    "Quantidade",
    "Data de Registro",
    "Técnico Responsável",
    "Status",
]

LARGURAS_COLUNAS = [32, 20, 12, 20, 24, 16]

COR_CABECALHO = "1E3A5F"  # mesma cor da topbar do site, pra manter identidade visual


def gerar_planilha_equipamentos(itens):
    """Recebe uma lista de objetos Equipamento e devolve um BytesIO com o .xlsx pronto
    para ser enviado via send_file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"

    ws.append(CABECALHOS)

    fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(
        start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid"
    )
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(CABECALHOS) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = alinhamento_cabecalho

    for item in itens:
        ws.append(
            [
                item.nome,
                item.categoria,
                item.quantidade,
                item.data_registro.strftime("%d/%m/%Y %H:%M"),
                item.tecnico_responsavel,
                item.status,
            ]
        )

    fonte_padrao = Font(name="Arial", size=11)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = fonte_padrao

    for idx, largura in enumerate(LARGURAS_COLUNAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = largura

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
